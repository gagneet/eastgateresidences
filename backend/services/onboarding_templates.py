# @featuretrace:onboarding — canonical CSV/XLSX template registry for every onboarding intake file.
# @featuretrace:financial-onboarding — the column contract every historical financial import is validated against.
# Layer: backend-service
# Data flow: this registry → GET /onboarding/templates[/{type}] → strata manager fills it in →
#            POST /onboarding/scheme/{session_id}/import-* → historical_* staging collections.
# Related: backend/routers/onboarding.py, backend/routers/financial_import.py,
#          frontend/src/pages/dashboard/admin/OnboardingWizard.jsx
"""Canonical column contract for every onboarding import file.

**Why this exists.** The onboarding wizard used to generate its own CSV templates
client-side from a hardcoded header string per step. Those strings had drifted
from what the endpoints actually parse — every one of them named columns the
backend never reads and omitted columns it requires — so a manager who downloaded
a template, filled it in and uploaded it got a 422 (or, worse, a silent import of
all-zero values). This module is the single source of truth: the same dict drives
the downloadable template AND is asserted against the endpoints' own
``required_columns`` by ``tests/backend/test_onboarding_templates.py``, so the two
cannot drift again.

Each entry declares:
  ``filename``   — download name (extension is swapped per format)
  ``endpoint``   — the import endpoint this file is uploaded to
  ``field``      — the multipart field name that endpoint expects for this file
  ``required``   — columns the endpoint rejects the upload without
  ``optional``   — columns the endpoint reads when present
  ``example``    — one illustrative row (values keyed by column)
  ``description``— one-line human summary shown in the wizard
"""
from __future__ import annotations

import csv
import io
from typing import Any

ONBOARDING_TEMPLATES: dict[str, dict[str, Any]] = {
    "lots": {
        "filename": "lots",
        "endpoint": "POST /onboarding/scheme/{session_id}/lots",
        "field": None,  # parsed client-side into JSON {lots: [...]}
        "description": "One row per lot/unit in the scheme, with its unit entitlement.",
        "required": ["lot_number"],
        "optional": ["unit_number", "lot_use", "floor_area_sqm", "entitlement_units", "owner_email"],
        "example": {
            "lot_number": "1", "unit_number": "101", "lot_use": "residential",
            "floor_area_sqm": "88.5", "entitlement_units": "115",
            "owner_email": "owner@example.com",
        },
    },
    "owner_transfers": {
        "filename": "owner_transfers",
        "endpoint": "POST /onboarding/scheme/{session_id}/import-owner-transfers",
        "field": "transfers_file",
        "description": "Historical change-of-ownership events (CSV 05) — becomes bitemporal ownership periods.",
        "required": ["lot_number", "effective_from", "new_owner_name", "previous_owner_name"],
        "optional": ["effective_to", "notes"],
        "example": {
            "lot_number": "1", "effective_from": "2023-04-01", "effective_to": "",
            "new_owner_name": "Jane Smith", "previous_owner_name": "John Smith", "notes": "",
        },
    },
    "quarterly_levies": {
        "filename": "quarterly_levies",
        "endpoint": "POST /onboarding/scheme/{session_id}/import-historical-financials",
        "field": "quarterly_levies",
        "description": "One row per (year, quarter, lot) levy issuance (CSV 06).",
        "required": ["year", "quarter", "lot_number", "admin_levy_amount", "sinking_levy_amount"],
        "optional": [
            "unit_number", "due_date", "period_start", "period_end", "uoe", "total_uoe",
            "total_levy_amount", "admin_levy_amount_excl_gst", "sinking_levy_amount_excl_gst", "notes",
        ],
        "example": {
            "year": "2025", "quarter": "1", "lot_number": "1", "unit_number": "101",
            "admin_levy_amount": "980.53", "sinking_levy_amount": "286.19",
            "due_date": "2025-03-31", "period_start": "2025-01-01", "period_end": "2025-03-31",
            "uoe": "115", "total_uoe": "10000", "total_levy_amount": "1266.72",
            "admin_levy_amount_excl_gst": "891.39", "sinking_levy_amount_excl_gst": "260.17", "notes": "",
        },
    },
    "admin_fund_summary": {
        "filename": "admin_fund_summary",
        "endpoint": "POST /onboarding/scheme/{session_id}/import-historical-financials",
        "field": "admin_fund_summary",
        "description": "One row per financial year of ADMIN fund totals from the audited statements (CSV 07).",
        "required": ["year", "levy_income"],
        "optional": [
            "period_start", "period_end", "budget_proposed", "other_income", "total_income",
            "total_expenses_audited", "surplus_deficit", "opening_balance", "closing_balance",
            "reconciliation_note", "notes",
        ],
        "example": {
            "year": "2025", "levy_income": "340870.20", "period_start": "2025-01-01",
            "period_end": "2025-12-31", "budget_proposed": "340870.20", "other_income": "1250.00",
            "total_income": "342120.20", "total_expenses_audited": "331004.55",
            "surplus_deficit": "11115.65", "opening_balance": "48210.11",
            "closing_balance": "59325.76", "reconciliation_note": "", "notes": "",
        },
    },
    "sinking_fund_summary": {
        "filename": "sinking_fund_summary",
        "endpoint": "POST /onboarding/scheme/{session_id}/import-historical-financials",
        "field": "sinking_fund_summary",
        "description": "One row per financial year of SINKING/capital-works fund totals (CSV 08).",
        "required": ["year", "levy_income"],
        "optional": [
            "period_start", "period_end", "budget_proposed", "other_income", "total_income",
            "total_expenses_audited", "surplus_deficit", "opening_balance", "closing_balance",
            "reconciliation_note", "notes",
        ],
        "example": {
            "year": "2025", "levy_income": "99504.90", "period_start": "2025-01-01",
            "period_end": "2025-12-31", "budget_proposed": "99504.90", "other_income": "0",
            "total_income": "99504.90", "total_expenses_audited": "61220.00",
            "surplus_deficit": "38284.90", "opening_balance": "120880.00",
            "closing_balance": "159164.90", "reconciliation_note": "", "notes": "",
        },
    },
    "arrears": {
        "filename": "arrears",
        "endpoint": "POST /onboarding/scheme/{session_id}/import-historical-financials",
        "field": "arrears",
        "description": "Per-lot arrears snapshot at the end of the last closed financial year (CSV 09).",
        "required": ["lot_number", "admin_arrears", "sinking_arrears"],
        "optional": [
            "unit_number", "total_arrears", "admin_closing_balance",
            "sinking_closing_balance", "data_source", "notes",
        ],
        "example": {
            "lot_number": "1", "unit_number": "101", "admin_arrears": "0",
            "sinking_arrears": "0", "total_arrears": "0", "admin_closing_balance": "0",
            "sinking_closing_balance": "0", "data_source": "audited_statement", "notes": "",
        },
    },
    "outstanding": {
        "filename": "outstanding",
        "endpoint": "POST /onboarding/scheme/{session_id}/import-historical-financials",
        "field": "outstanding",
        "description": "Per-lot outstanding balance as at the cutover date (CSV 10).",
        "required": ["lot_number", "admin_outstanding", "sinking_outstanding"],
        "optional": [
            "unit_number", "as_at_date", "total_outstanding", "data_source", "notes",
        ],
        "example": {
            "lot_number": "1", "unit_number": "101", "as_at_date": "2026-06-01",
            "admin_outstanding": "1241.22", "sinking_outstanding": "362.19",
            "total_outstanding": "1603.41", "data_source": "portal_snapshot", "notes": "",
        },
    },
    "opening_balances": {
        "filename": "opening_balances",
        "endpoint": "POST /onboarding/scheme/{session_id}/import-opening-balances",
        "field": "opening_balances_file",
        "description": "Fund opening balances at cutover, one row per fund (CSV 11).",
        "required": ["fund_type", "opening_balance_amount"],
        "optional": ["as_at_date", "bsb", "account_number", "account_name", "evidence_source", "notes"],
        "example": {
            "fund_type": "admin", "opening_balance_amount": "59325.76", "as_at_date": "2026-01-01",
            "bsb": "062-000", "account_number": "12345678", "account_name": "OC Admin Fund",
            "evidence_source": "bank_statement", "notes": "",
        },
    },
    "historical_expenses": {
        "filename": "historical_expenses",
        "endpoint": "POST /onboarding/scheme/{session_id}/import-historical-expenses",
        "field": "expenses",
        "description": "Settled historical expense transactions from AGM records/audited reports/invoices.",
        "required": ["vendor_name", "category_name", "financial_year", "amount_ex_gst"],
        "optional": [
            "invoice_number", "fund_type", "transaction_date", "gst_amount",
            "description", "evidence_references", "derivation_level",
        ],
        "example": {
            "vendor_name": "ACME Lifts Pty Ltd", "invoice_number": "INV-2025-0412",
            "category_name": "Lift Maintenance", "fund_type": "admin", "financial_year": "2025",
            "transaction_date": "2025-08-14", "amount_ex_gst": "2400.00", "gst_amount": "240.00",
            "description": "Quarterly lift service", "evidence_references": "",
            "derivation_level": "exact",
        },
    },
}

# Files that must be uploaded together in one request, in endpoint order.
HISTORICAL_FINANCIALS_FILE_GROUP: list[str] = [
    "quarterly_levies", "admin_fund_summary", "sinking_fund_summary", "arrears", "outstanding",
]


def template_columns(template_type: str) -> list[str]:
    """Ordered column list (required first, then optional) for a template."""
    spec = ONBOARDING_TEMPLATES[template_type]
    return [*spec["required"], *spec["optional"]]


def render_template_csv(template_type: str) -> bytes:
    """Render a template as UTF-8 CSV bytes: header row + one example row."""
    spec = ONBOARDING_TEMPLATES[template_type]
    columns = template_columns(template_type)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    writer.writerow({c: spec["example"].get(c, "") for c in columns})
    return buf.getvalue().encode("utf-8")


def render_template_xlsx(template_type: str) -> bytes:
    """Render a template as an .xlsx workbook: bolded header row + one example row.

    Required columns are visually distinguished (bold + tinted) so a manager
    filling the sheet in can see at a glance which ones the import will reject
    the file without.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    spec = ONBOARDING_TEMPLATES[template_type]
    columns = template_columns(template_type)
    required = set(spec["required"])

    wb = Workbook()
    ws = wb.active
    ws.title = template_type[:31]

    required_fill = PatternFill("solid", start_color="FFF2CC8F", end_color="FFF2CC8F")
    optional_fill = PatternFill("solid", start_color="FFE8E8E8", end_color="FFE8E8E8")

    ws.append(columns)
    for idx, column in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True)
        cell.fill = required_fill if column in required else optional_fill
        cell.alignment = Alignment(horizontal="left")
        ws.column_dimensions[cell.column_letter].width = max(14, min(32, len(column) + 4))
    ws.append([spec["example"].get(c, "") for c in columns])
    ws.freeze_panes = "A2"

    notes = wb.create_sheet("README")
    notes["A1"] = f"Template: {template_type}"
    notes["A2"] = spec["description"]
    notes["A3"] = f"Upload to: {spec['endpoint']}"
    notes["A4"] = "Highlighted (amber) columns are REQUIRED — the import is rejected without them."
    notes["A5"] = "Grey columns are optional; leave blank if unknown. Do not rename or reorder columns."
    notes["A6"] = "Delete the example row before uploading. Amounts are plain numbers (no $ or thousands separators)."
    notes["A7"] = "Dates are YYYY-MM-DD."
    notes.column_dimensions["A"].width = 110

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
