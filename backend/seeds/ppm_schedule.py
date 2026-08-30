"""
Seed script: Preventive Maintenance Plan (PPM) schedule.

Source: docs/finances/EastGate-UP13195-PPM-January2026.xlsx
Building ID: 13195 — East Gate Residences, 14 Hoolihan Street, Denman Prospect

Usage (from backend/ directory):
    python3 seeds/ppm_schedule.py

Idempotent: upserts keyed on (building_id, description, section).
Expected: 79 items, 15 sections, 8 compliance items.
"""

import sys
import uuid
from datetime import datetime, date, timezone, timedelta
from pathlib import Path

import asyncio
import logging

# Allow running from any location — add backend/ to path
_SCRIPT_DIR = Path(__file__).resolve().parent
_BACKEND_DIR = _SCRIPT_DIR.parent
if str(_BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(_BACKEND_DIR))

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

from database import db
from pymongo import UpdateOne

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

BUILDING_ID = "13195"
EXCEL_PATH = Path(__file__).resolve().parents[2] / "docs" / "finances" / "EastGate-UP13195-PPM-January2026.xlsx"

# ─── Frequency normalisation ──────────────────────────────────────────────────

_FREQ_MAP = {
    "annually": ("annual", 365),
    "annual": ("annual", 365),
    "quarterly": ("quarterly", 90),
    "3 monthly": ("quarterly", 90),
    "per agreement (quarterly)": ("quarterly", 90),
    "per agreement": ("annual", 365),
    "monthly": ("monthly", 30),
    "6 monthly": ("6monthly", 180),
    "6 Monthly": ("6monthly", 180),
    "6mthly - annually": ("6monthly", 180),
    "biannually": ("biannual", 180),
    "biannanual": ("biannual", 180),
    "5 yearly": ("5yearly", 1825),
    "one off": ("annual", 365),  # one-off items default to annual cadence
    "as required": ("annual", 365),
}


def _normalise_frequency(raw) -> tuple[str, int]:
    """Return (normalised_name, days) for a raw frequency cell value."""
    if raw is None:
        return ("annual", 365)
    # datetime object in frequency cell → treat as one-off / annual
    if isinstance(raw, (datetime, date)):
        return ("annual", 365)
    s = str(raw).strip().lower()
    for key, val in _FREQ_MAP.items():
        if s == key.lower():
            return val
    # Fallback
    return ("annual", 365)


# ─── Date parsing helpers ─────────────────────────────────────────────────────

def _parse_date(cell_val) -> str | None:
    """Return ISO date string (YYYY-MM-DD) or None if unparseable."""
    if cell_val is None:
        return None
    if isinstance(cell_val, (datetime, date)):
        if isinstance(cell_val, datetime):
            return cell_val.date().isoformat()
        return cell_val.isoformat()
    s = str(cell_val).strip()
    if not s or s == "None":
        return None
    # Try a few formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%y", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _compute_next_due(last_inspected_str: str | None, frequency_days: int) -> str:
    """Compute next_due_date = last_inspected + frequency_days or today + frequency_days."""
    today = date.today()
    if last_inspected_str:
        try:
            base = date.fromisoformat(last_inspected_str)
            return (base + timedelta(days=frequency_days)).isoformat()
        except ValueError:
            pass
    return (today + timedelta(days=frequency_days)).isoformat()


def _compute_status(next_due_str: str) -> str:
    """Generated function header.

    Function: _compute_status
    Path: backend/seeds/ppm_schedule.py

    Note: Generated inventory header. Replace or expand this with reviewed business-purpose documentation before relying on it as source commentary.
    """
    today = date.today()
    try:
        next_due = date.fromisoformat(next_due_str)
        delta = (next_due - today).days
        if delta < 0:
            return "overdue"
        elif delta <= 30:
            return "due_soon"
        return "scheduled"
    except ValueError:
        return "scheduled"


# ─── Excel parsing ────────────────────────────────────────────────────────────

YEAR_COLUMNS = list(range(2024, 2039))  # columns 10–24 (0-indexed) = years 2024–2038

# Non-numeric values that appear in the lifespan column
_NON_NUMERIC_LIFESPANS = {"", "None", "Ad Hoc", "As required", "July 25 expected"}


def _parse_lifespan(cell_val) -> int | None:
    """Parse the lifespan cell value to an integer number of years, or None."""
    if cell_val is None:
        return None
    if isinstance(cell_val, (int, float)):
        return int(cell_val) if not isinstance(cell_val, float) or cell_val == int(cell_val) else int(cell_val)
    s = str(cell_val).strip()
    if s in _NON_NUMERIC_LIFESPANS:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _parse_excel(path: Path) -> list[dict]:
    """Parse the PPM Excel file and return a list of schedule item dicts."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active

    items: list[dict] = []
    current_section: str = "GENERAL"
    now = datetime.now(timezone.utc).isoformat()

    for r in range(10, ws.max_row + 1):
        row_data = [ws.cell(row=r, column=c).value for c in range(1, 26)]

        col0 = row_data[0]  # description
        col4 = row_data[4]  # inspection_type
        col6 = row_data[6]  # frequency

        # Skip empty rows
        if col0 is None:
            continue

        # Section header: col[4] and col[6] are both None
        if col4 is None and col6 is None:
            # Stop at inspection type legend rows
            if str(col0).strip().lower().startswith("inspection type"):
                break
            current_section = str(col0).strip()
            continue

        # Data row — must have col[6] (frequency) OR col[4] (inspection_type)
        if col4 is None and col6 is None:
            continue  # skip if both missing (already handled above)

        description = str(col0).strip()
        vendor_name = str(row_data[1]).strip() if row_data[1] else None
        identifier = str(row_data[2]).strip() if row_data[2] else None
        install_date = _parse_date(row_data[3])
        inspection_type = str(col4).strip() if col4 else None
        standard = str(row_data[5]).strip() if row_data[5] else None
        last_inspected_raw = row_data[7]
        last_inspected = _parse_date(last_inspected_raw)
        inspector = str(row_data[8]).strip() if row_data[8] else None
        lifespan_raw = row_data[9]
        lifespan_years = _parse_lifespan(lifespan_raw)

        frequency, frequency_days = _normalise_frequency(col6)
        next_due_date = _compute_next_due(last_inspected, frequency_days)
        status = _compute_status(next_due_date)
        is_compliance = bool(inspection_type and "Compliance" in inspection_type)

        # Capital expenditure years: X in columns 10–24
        capital_years: list[int] = []
        for idx, year in enumerate(YEAR_COLUMNS):
            cell_val = row_data[10 + idx]
            if cell_val is not None and str(cell_val).strip().upper() == "X":
                capital_years.append(year)

        # Determine event colour
        if is_compliance:
            color = "red"
        elif inspector and "Licenced Contractor" in inspector:
            color = "orange"
        elif date.today().year in capital_years:
            color = "purple"
        else:
            color = "blue"

        schedule_id = str(uuid.uuid4())

        items.append(
            {
                "schedule_id": schedule_id,
                "building_id": BUILDING_ID,
                "section": current_section,
                "description": description,
                "vendor_name": vendor_name,
                "vendor_contact": None,
                "contractor_id": None,
                "identifier": identifier,
                "install_date": install_date,
                "inspection_type": inspection_type,
                "standard": standard,
                "frequency": frequency,
                "frequency_days": frequency_days,
                "last_completed_date": last_inspected,
                "next_due_date": next_due_date,
                "status": status,
                "is_compliance": is_compliance,
                "lifespan_years": lifespan_years,
                "capital_years": capital_years,
                "inspector": inspector,
                "notification_log": [],
                "completion_log": [],
                "color": color,
                "created_at": now,
                "updated_at": now,
            }
        )

    return items


# ─── Seeding ──────────────────────────────────────────────────────────────────

async def _create_calendar_events(items: list[dict]) -> None:
    """Upsert calendar events for all seeded PPM items."""
    now = datetime.now(timezone.utc).isoformat()
    ops = []
    for item in items:
        section = item.get("section", "")
        description = item.get("description", "")
        title_raw = f"{section}: {description}"
        title = title_raw[:80]

        event_doc = {
            "id": str(uuid.uuid4()),
            "building_id": BUILDING_ID,
            "event_type": "ppm",
            "title": title,
            "description": description,
            "start_date": item["next_due_date"],
            "end_date": None,
            "is_public": False,
            "is_recurring": True,
            "ppm_schedule_id": item["schedule_id"],
            "metadata": {"color": item.get("color", "blue"), "section": section},
            "source": "ppm",
            "created_by": "system",
            "created_at": now,
        }
        ops.append(
            UpdateOne(
                {
                    "building_id": BUILDING_ID,
                    "event_type": "ppm",
                    "ppm_schedule_id": item["schedule_id"],
                    "completed_at": {"$exists": False},
                },
                {"$setOnInsert": event_doc},
                upsert=True,
            )
        )

    if ops:
        await db.events.bulk_write(ops, ordered=False)
        logger.info("Calendar events upserted: %d", len(ops))


async def seed_ppm() -> None:
    """Parse Excel and upsert all PPM items. Idempotent."""
    if not EXCEL_PATH.exists():
        logger.error("Excel file not found: %s", EXCEL_PATH)
        sys.exit(1)

    logger.info("Parsing Excel: %s", EXCEL_PATH)
    items = _parse_excel(EXCEL_PATH)
    logger.info("Parsed %d items from Excel", len(items))

    if not items:
        logger.warning("No items parsed — check Excel format.")
        return

    # Build bulk upsert operations keyed on (building_id, description, section)
    ops = []
    now = datetime.now(timezone.utc).isoformat()

    for item in items:
        filter_key = {
            "building_id": item["building_id"],
            "description": item["description"],
            "section": item["section"],
        }
        # On insert: set all fields including schedule_id (new UUID)
        # On update: preserve existing schedule_id, update mutable fields
        ops.append(
            UpdateOne(
                filter_key,
                {
                    "$setOnInsert": {
                        "schedule_id": item["schedule_id"],
                        "id": item["schedule_id"],
                        "created_at": item["created_at"],
                    },
                    "$set": {
                        "building_id": item["building_id"],
                        "section": item["section"],
                        "description": item["description"],
                        "vendor_name": item["vendor_name"],
                        "vendor_contact": item.get("vendor_contact"),
                        "contractor_id": item.get("contractor_id"),
                        "identifier": item.get("identifier"),
                        "install_date": item.get("install_date"),
                        "inspection_type": item["inspection_type"],
                        "standard": item["standard"],
                        "frequency": item["frequency"],
                        "frequency_days": item["frequency_days"],
                        "last_completed_date": item["last_completed_date"],
                        "next_due_date": item["next_due_date"],
                        "status": item["status"],
                        "is_compliance": item["is_compliance"],
                        "lifespan_years": item["lifespan_years"],
                        "capital_years": item["capital_years"],
                        "inspector": item.get("inspector"),
                        "color": item["color"],
                        "updated_at": now,
                    },
                },
                upsert=True,
            )
        )

    result = await db.maintenance_schedules.bulk_write(ops, ordered=False)
    logger.info(
        "Upsert complete — inserted: %d, modified: %d, matched: %d",
        result.upserted_count,
        result.modified_count,
        result.matched_count,
    )

    # Fetch all seeded items (with their actual schedule_ids from DB)
    seeded = await db.maintenance_schedules.find(
        {"building_id": BUILDING_ID}, {"_id": 0}
    ).to_list(length=500)

    await _create_calendar_events(seeded)

    # Print summary
    total = len(seeded)
    sections = len({i["section"] for i in seeded})
    compliance = sum(1 for i in seeded if i.get("is_compliance"))
    overdue = sum(1 for i in seeded if i.get("status") == "overdue")
    due_soon = sum(1 for i in seeded if i.get("status") == "due_soon")

    print(f"\n{'=' * 50}")
    print(f"PPM Seed Summary")
    print(f"{'=' * 50}")
    print(f"  Total items:      {total}")
    print(f"  Sections:         {sections}")
    print(f"  Compliance items: {compliance}")
    print(f"  Overdue:          {overdue}")
    print(f"  Due soon:         {due_soon}")
    print(f"{'=' * 50}\n")


if __name__ == "__main__":
    asyncio.run(seed_ppm())
