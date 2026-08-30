# @featuretrace:onboarding — shared CSV/XLSX tabular-upload parsing boundary.
# @featuretrace:financial-onboarding — CSV/XLSX intake boundary for the financial-import path.
# Layer: backend-util
# Data flow: UploadFile bytes → parse_tabular_bytes() → list[dict] rows → onboarding /
#            financial-import row processors.
# Related: backend/routers/onboarding.py, backend/routers/financial_import.py,
#          backend/services/financial_import_service.py, backend/utils/file_scan.py
"""Tabular (CSV **and** XLSX) upload parsing — one boundary, one row shape.

Every onboarding / financial-import upload path used to be CSV-only: the declared
content-type allowlist rejected the real ``.xlsx`` MIME type, the Magika ``csv``
scan context rejected the detected label, and the parser was a bare
``csv.DictReader(content.decode("utf-8"))`` that raises ``UnicodeDecodeError`` on
any binary workbook.  Strata managers export from Excel far more often than they
hand-write CSV, so this module makes both formats enter through the *same* parse
boundary and produce the *same* ``list[dict]`` of string-valued rows — the
downstream row processors are unchanged and stay format-agnostic.

Contract
--------
* Returns ``(headers, rows)`` where every value is a ``str`` (never ``None``,
  never a float/datetime), because the existing CSV row processors call
  ``_safe_float`` / ``.strip()`` on raw ``csv.DictReader`` output.
* XLSX numbers are emitted without a trailing ``.0`` when integral, so
  ``lot_number`` reads ``"7"`` rather than ``"7.0"`` — that string is used as a
  Mongo/Postgres lookup key and ``"7.0"`` would silently miss every lot.
* Dates/datetimes are emitted ISO-8601 (``YYYY-MM-DD``), matching what the CSV
  templates carry.
* Legacy binary ``.xls`` is explicitly NOT supported (no reader in
  ``requirements.txt``); it raises a clear 415 telling the user to re-save.
"""
from __future__ import annotations

import csv
import io
import zipfile
from datetime import date, datetime
from typing import Any

from fastapi import HTTPException

# Content types accepted for a tabular upload.
# NOTE: "application/vnd.ms-excel" is what Windows browsers frequently send for a
# plain .csv when Excel is the registered handler — it is NOT proof of a binary
# .xls.  Actual format detection is done from the bytes, not this header.
TABULAR_CONTENT_TYPES: frozenset[str] = frozenset({
    "text/csv",
    "application/csv",
    "text/plain",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
})

_ZIP_MAGIC = b"PK\x03\x04"
_OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # legacy .xls / OLE2 compound file

# Guard against a zip bomb: an .xlsx is a zip, and openpyxl will happily inflate
# whatever it is given.  10 MB compressed (the upload cap) can inflate to GBs.
_MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024


def looks_like_xlsx(content: bytes) -> bool:
    """True when the bytes are an OOXML workbook (not merely any zip)."""
    if not content.startswith(_ZIP_MAGIC):
        return False
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            return any(n.startswith("xl/") for n in zf.namelist())
    except zipfile.BadZipFile:
        return False


def looks_like_legacy_xls(content: bytes) -> bool:
    """True for the legacy binary .xls (OLE2) format, which we cannot read."""
    return content.startswith(_OLE_MAGIC)


def assert_content_type_allowed(content_type: str | None, filename: str) -> None:
    """Raise 415 when the declared content-type is outside the tabular allowlist."""
    ct = (content_type or "").split(";")[0].strip().lower()
    if ct and ct not in TABULAR_CONTENT_TYPES:
        raise HTTPException(
            status_code=415,
            detail=(
                f"Unsupported file type '{ct}' for '{filename}'. "
                "Please upload a CSV (.csv) or Excel workbook (.xlsx)."
            ),
        )


def _cell_to_str(value: Any) -> str:
    """Normalise an openpyxl cell value to the string a CSV would have carried."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        # Midnight datetimes are date-only cells in Excel's model.
        if value.hour == value.minute == value.second == 0:
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        # 7.0 -> "7": these strings are used as lot/unit lookup keys.
        return str(int(value)) if value.is_integer() else repr(value)
    return str(value).strip()


def _decode_text(content: bytes) -> str:
    """Decode CSV bytes, handling a UTF-8 BOM and falling back to latin-1."""
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError:
        return content.decode("latin-1")


def _parse_xlsx(content: bytes, filename: str) -> tuple[list[str], list[dict]]:
    """Parse the first worksheet of an .xlsx into (headers, rows)."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - openpyxl is a hard requirement
        raise HTTPException(
            status_code=415,
            detail="Excel (.xlsx) upload is unavailable on this server; please upload a CSV instead.",
        ) from exc

    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        total = sum(info.file_size for info in zf.infolist())
    if total > _MAX_UNCOMPRESSED_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"'{filename}' expands to {total // (1024 * 1024)} MB, which exceeds the workbook limit.",
        )

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=422,
            detail=f"'{filename}' could not be read as an Excel workbook: {exc}",
        ) from exc

    try:
        ws = wb[wb.sheetnames[0]] if wb.sheetnames else None
        if ws is None:
            return [], []

        row_iter = ws.iter_rows(values_only=True)
        headers: list[str] = []
        for raw in row_iter:
            candidate = [_cell_to_str(c) for c in raw]
            if any(c for c in candidate):
                headers = [c for c in candidate]
                break
        if not headers:
            return [], []
        # Trim trailing all-empty header columns (Excel pads the used range).
        while headers and not headers[-1]:
            headers.pop()

        rows: list[dict] = []
        for raw in row_iter:
            values = [_cell_to_str(c) for c in raw]
            if not any(v for v in values):
                continue  # skip blank spacer rows
            rows.append({h: (values[i] if i < len(values) else "") for i, h in enumerate(headers) if h})
        return [h for h in headers if h], rows
    finally:
        wb.close()


def parse_tabular_bytes(content: bytes, *, filename: str = "") -> tuple[list[str], list[dict]]:
    """Parse CSV **or** XLSX bytes into ``(headers, rows)`` of plain strings.

    Format is detected from the bytes (magic number), never from the filename or
    the browser-declared content-type — both of which are routinely wrong.
    """
    if looks_like_legacy_xls(content):
        raise HTTPException(
            status_code=415,
            detail=(
                f"'{filename or 'file'}' is a legacy Excel 97-2003 workbook (.xls), which is not "
                "supported. Re-save it as .xlsx or .csv and upload again."
            ),
        )
    if looks_like_xlsx(content):
        return _parse_xlsx(content, filename or "workbook.xlsx")

    reader = csv.DictReader(io.StringIO(_decode_text(content)))
    headers = list(reader.fieldnames or [])
    return headers, list(reader)


def assert_required_columns(
        headers: list[str], required_columns: list[str], filename: str,
) -> None:
    """Raise 422 listing every missing required column.

    Without this check a file with the wrong schema parses without error and every
    mismatched field lookup silently coerces to 0.0/None — producing wrong imported
    data with no indication anything went wrong.
    """
    actual = set(headers)
    missing = [c for c in required_columns if c not in actual]
    if missing:
        raise HTTPException(
            status_code=422,
            detail=(
                f"'{filename}' is missing required column(s) {missing}. "
                f"Found columns: {sorted(actual)}. This file does not match the "
                "expected schema for this import — check you're uploading the correct file."
            ),
        )


def tabular_bytes_to_csv_bytes(content: bytes, *, filename: str = "") -> bytes:
    """Return CSV bytes for either CSV or XLSX input.

    CSV input is returned unchanged (byte-for-byte — no re-encoding, so quoting
    and encoding quirks the existing processors already tolerate are preserved).
    XLSX input is parsed and re-emitted as UTF-8 CSV. This lets an upload
    boundary accept both formats while every downstream ``process_*_csv(content:
    bytes)`` row processor keeps its existing contract.
    """
    if not looks_like_legacy_xls(content) and not looks_like_xlsx(content):
        return content
    headers, rows = parse_tabular_bytes(content, filename=filename)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({h: row.get(h, "") for h in headers})
    return buf.getvalue().encode("utf-8")
